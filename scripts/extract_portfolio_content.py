import json
import os
import shutil
import subprocess
from pathlib import Path

from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OUT = ROOT / "src" / "data" / "portfolio-content.json"
ASSET_ROOT = ROOT / "public" / "portfolio-assets" / "extracted"
ASSET_URL = "/portfolio-assets/extracted"
POPPLER_BIN = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "bin"
    / "override"
)
PDFTOPPM = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "native"
    / "poppler"
    / "Library"
    / "bin"
    / "pdftoppm.exe"
)


def clean(value):
    return " ".join(str(value).replace("\u3000", " ").split())


def content_type_ext(content_type, fallback):
    mapping = {
        "image/jpeg": ".jpg",
        "image/png": ".png",
        "image/gif": ".gif",
        "image/tiff": ".tif",
        "image/bmp": ".bmp",
        "image/webp": ".webp",
    }
    return mapping.get(content_type, fallback or ".bin")


def reset_asset_dir(slug):
    target = ASSET_ROOT / slug
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)
    return target


def iter_block_items(doc):
    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn("w:p"):
            yield Paragraph(child, doc)
        elif child.tag == qn("w:tbl"):
            yield Table(child, doc)


def paragraph_image_rids(paragraph):
    rids = []
    for blip in paragraph._element.xpath(".//*[local-name()='blip']"):
        rid = blip.get(qn("r:embed")) or blip.get(qn("r:link"))
        if rid:
            rids.append(rid)
    return rids


def save_docx_image(doc, rid, asset_dir, slug, index, saved):
    if rid in saved:
        return saved[rid]

    rel = doc.part.rels[rid]
    part = rel.target_part
    ext = content_type_ext(getattr(part, "content_type", ""), Path(part.partname).suffix)
    filename = f"image-{index:03d}{ext}"
    out = asset_dir / filename
    out.write_bytes(part.blob)
    src = f"{ASSET_URL}/{slug}/{filename}"
    saved[rid] = src
    return src


def table_rows(table):
    rows = []
    for row in table.rows:
        values = [clean(cell.text) for cell in row.cells]
        if any(values):
            rows.append(values)
    return rows


def run_size(run):
    if run.font.size:
        return run.font.size.pt
    rpr = run._element.rPr
    if rpr is not None and rpr.sz is not None and rpr.sz.val:
        try:
            return int(rpr.sz.val) / 2
        except (TypeError, ValueError):
            return None
    return None


def paragraph_sizes(paragraph):
    return [
        size
        for run in paragraph.runs
        if run.text.strip()
        for size in [run_size(run)]
        if size is not None
    ]


def paragraph_role(paragraph):
    style = paragraph.style
    style_name = (style.name if style else "").lower()
    if style_name.startswith("heading 1"):
        return "heading"
    if style_name.startswith("heading"):
        return "subheading"

    sizes = paragraph_sizes(paragraph)
    style_size = style.font.size.pt if style and style.font.size else None
    effective_size = max(sizes) if sizes else style_size
    if effective_size is not None and effective_size <= 10:
        return "caption"
    return "paragraph"


def docx_blocks(path, slug):
    doc = Document(path)
    asset_dir = reset_asset_dir(slug)
    saved_images = {}
    image_index = 1
    blocks = []

    for item in iter_block_items(doc):
        if isinstance(item, Paragraph):
            text = clean(item.text)
            if text:
                blocks.append(
                    {
                        "type": "paragraph",
                        "text": text,
                        "role": paragraph_role(item),
                    }
                )

            for rid in paragraph_image_rids(item):
                src = save_docx_image(doc, rid, asset_dir, slug, image_index, saved_images)
                image_index += 1
                blocks.append({"type": "image", "src": src})
        elif isinstance(item, Table):
            rows = table_rows(item)
            if rows:
                blocks.append(
                    {
                        "type": "table",
                        "headers": rows[0],
                        "rows": rows[1:],
                    }
                )

    return blocks


def pdf_blocks(path, slug):
    asset_dir = reset_asset_dir(slug)
    prefix = asset_dir / "page"
    env = os.environ.copy()
    env["PATH"] = str(POPPLER_BIN) + os.pathsep + env.get("PATH", "")
    subprocess.run(
        [str(PDFTOPPM), "-png", "-r", "145", str(path), str(prefix)],
        check=True,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )

    pages = sorted(
        asset_dir.glob("page-*.png"),
        key=lambda item: int(item.stem.split("-")[-1]),
    )
    return [
        {
            "type": "pdf-page",
            "src": f"{ASSET_URL}/{slug}/{page.name}",
            "caption": f"第 {index} 页",
        }
        for index, page in enumerate(pages, 1)
    ]


def block_texts(blocks):
    texts = []
    for block in blocks:
        if block["type"] == "paragraph":
            texts.append(block["text"])
        elif block["type"] == "table":
            texts.extend(" | ".join(row) for row in [block["headers"], *block["rows"]])
    return texts


def first_image(blocks):
    for block in blocks:
        if block["type"] in {"image", "pdf-page"}:
            return block["src"]
    return None


def xlsx_rows(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = []
    current_platform = ""
    current_genre = ""
    for row in sheet.iter_rows(values_only=True):
        values = [clean(value) if value is not None else "" for value in row[:4]]
        if not any(values):
            continue
        if values[0] and values[0] != "平台":
            current_platform = values[0]
        if values[1] and values[1] != "游戏类型":
            current_genre = values[1]
        if values[2] and values[2] != "游戏名称":
            rows.append([current_platform, current_genre, values[2], values[3]])
    return rows


categories = [
    {
        "slug": "game-design",
        "title": "AI 游戏策划",
        "kicker": "GAME DESIGN",
        "description": "互动视频方向的游戏概念、玩法循环、视觉基调与叙事机制。",
    },
    {
        "slug": "game-research",
        "title": "游戏拆解与研究",
        "kicker": "SYSTEM ANALYSIS",
        "description": "战斗系统、社交设计、玩家行为与机制引导拆解。",
    },
    {
        "slug": "articles",
        "title": "信号文章",
        "kicker": "CULTURE WRITING",
        "description": "IP 改编、角色设计、社群运营、互联网文化与漫画表达观察。",
    },
    {
        "slug": "scripts",
        "title": "原创剧本",
        "kicker": "SCRIPT WRITING",
        "description": "短片与剧作构思，关注校园压力、童年英雄、城市更新与乡土记忆。",
    },
    {
        "slug": "reviews",
        "title": "影评 & 剧评",
        "kicker": "CRITICISM",
        "description": "电影语言、戏剧语言、叙事结构与泛文娱榜单分析。",
    },
    {
        "slug": "experience",
        "title": "游戏经历",
        "kicker": "PLAY HISTORY",
        "description": "长期游玩记录与跨类型体验样本。",
    },
]


work_specs = [
    ("game-design", "dwarf-lost-in-the-dark-forest", "DWARF: Lost in the Dark Forest", "互动视频 / 恐怖探索", DOCS / "AI游戏策划" / "DWARF" / "dwarf-lost-in-the-dark-forest-gdd.docx", "docx", ["互动视频", "恐怖探索", "Low Poly", "Jump Scare"], "/portfolio-assets/games/dwarf-cover.png"),
    ("game-design", "stones-feast", "STONE'S FEAST", "单人互动恐怖整蛊游戏", DOCS / "AI游戏策划" / "STONE'S FEAST" / "stones-feast-game-design.docx", "docx", ["黑色幽默", "像素恐怖", "道具组合", "循环叙事"], "/portfolio-assets/games/stones-feast-cover.png"),
    ("game-design", "the-cold-trial", "The Cold Trial", "互动视频推理 / 法庭叙事", DOCS / "AI游戏策划" / "TheColdTrial" / "the-cold-trial-gdd.docx", "docx", ["互动推理", "档案拼图", "历史案件", "道德困境"], "/portfolio-assets/games/cold-trial-cover.png"),
    ("game-research", "peak-social-design", "游戏 PEAK 社交设计洞察", "玩家画像 / 机制引导 / 互动行为", DOCS / "游戏peak社交设计洞察.pdf", "pdf", ["社交设计", "玩家画像", "联机体验", "机制引导"], None),
    ("game-research", "rock-kingdom-world-breakdown", "洛克王国：世界拆解", "战斗系统 / 回合制 / 资源博弈", DOCS / "游戏拆解" / "洛克王国：世界拆解.pdf", "pdf", ["系统拆解", "战斗机制", "资源管理", "应对关系"], None),
    ("articles", "transformers-ip-games", "以《变形金刚》为例洞察IP改编游戏兴衰成因", "IP 改编 / 游戏化适配", DOCS / "信号文章" / "以《变形金刚》为例洞察IP改编游戏兴衰成因.docx", "docx", ["IP改编", "变形金刚", "游戏化", "案例分析"], None),
    ("articles", "rock-kingdom-farewell", "关于洛克王国，关于离别的童话", "叙事体验 / 情绪主题", DOCS / "信号文章" / "关于洛克王国，关于离别的童话.docx", "docx", ["洛克王国", "叙事体验", "离别", "情绪设计"], None),
    ("articles", "pokemon-character-design", "宝可梦形象设计思路分析", "角色设计 / 造型逻辑", DOCS / "信号文章" / "宝可梦形象设计思路分析.docx", "docx", ["宝可梦", "角色设计", "形象分析", "怪异美学"], None),
    ("articles", "helldivers2-community-ops", "绝地潜兵2社媒账号运营如何靠搞抽象活跃社群", "社群运营 / 动态叙事", DOCS / "信号文章" / "绝地潜兵2社媒账号运营如何靠搞抽象活跃社群.docx", "docx", ["社群运营", "动态叙事", "绝地潜兵2", "传播"], None),
    ("articles", "alien-meme-culture", "荒诞又不知“何意味”的大头外星人如何火出圈", "互联网文化 / 梗传播", DOCS / "信号文章" / "荒诞又不知“何意味”的大头外星人如何火出圈.docx", "docx", ["互联网文化", "梗图", "外星人", "传播机制"], None),
    ("articles", "chainsaw-man-dream", "说真的，《电锯人》的梦早该结束了。", "漫画评论 / 作者与受众", DOCS / "信号文章" / "说真的，《电锯人》的梦早该结束了。.docx", "docx", ["电锯人", "漫画评论", "受众沟通", "角色塑造"], None),
    ("articles", "fujimoto-cinematic-sense", "谈藤本树作品中颇受追捧的“电影感”背后的生效机制", "影像感 / 漫画表达", DOCS / "信号文章" / "谈藤本树作品中颇受追捧的“电影感”背后的生效机制.docx", "docx", ["藤本树", "电影感", "漫画表达", "情绪连接"], None),
    ("scripts", "disappear", "消失", "校园 / 时间重复 / 压力逃离", DOCS / "剧本" / "消失 剧本.docx", "docx", ["校园", "时间重复", "短片", "心理压力"], None),
    ("scripts", "iron-superman", "铁超人安在", "短片剧本 / 童年英雄 / 创作困境", DOCS / "剧本" / "铁超人安在 剧本.docx", "docx", ["童年英雄", "寻找", "短片", "创作困境"], None),
    ("scripts", "rain-alley-pagoda-tree", "雨巷槐", "城市更新 / 乡土记忆 / 老树保护", DOCS / "剧本" / "雨巷槐 剧本.docx", "docx", ["城市更新", "乡土记忆", "老树保护", "人物群像"], None),
    ("reviews", "caines-mutiny-stage-review", "“话”变--谈《哗变》", "戏剧语言 / 文学改编", next((DOCS / "影评&剧评").glob("*哗变*.docx")), "docx", ["戏剧语言", "文学改编", "北京人艺", "文本分析"], None),
    ("reviews", "rango-review", "《兰戈》：一场波澜壮阔的隐喻", "动画电影 / 身份认同", DOCS / "影评&剧评" / "《兰戈》：一场波澜壮阔的隐喻.docx", "docx", ["动画电影", "身份认同", "隐喻", "西部片"], None),
    ("reviews", "twelve-angry-men-review", "为了对话这盘醋包的一部好饺子--谈《十二怒汉》", "对白电影 / 叙事张力", DOCS / "影评&剧评" / "为了对话这盘醋包的一部好饺子——谈《十二怒汉》.docx", "docx", ["对白电影", "叙事张力", "群像", "剧本结构"], None),
    ("reviews", "2001-space-odyssey-review", "寰宇洄游--评《2001太空漫游》", "电影语言 / 表现性", DOCS / "影评&剧评" / "寰宇洄游——评《2001太空漫游》.docx", "docx", ["电影语言", "科幻", "表现性", "视听表达"], None),
    ("reviews", "film-ranking-analysis", "电影榜单重点单品分析", "泛文娱榜单 / 数据观察", DOCS / "影评&剧评" / "电影榜单重点单品分析.docx", "docx", ["榜单分析", "泛文娱", "数据观察", "单品研究"], None),
]


def main():
    works = []
    for category, slug, title, kind, path, fmt, tags, image in work_specs:
        blocks = docx_blocks(path, slug) if fmt == "docx" else pdf_blocks(path, slug)
        paragraphs = block_texts(blocks)
        works.append(
            {
                "category": category,
                "slug": slug,
                "title": title,
                "kind": kind,
                "format": fmt,
                "tags": tags,
                "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                "image": image or first_image(blocks),
                "blocks": blocks,
                "paragraphs": paragraphs,
                "paragraphCount": sum(1 for block in blocks if block["type"] == "paragraph"),
                "imageCount": sum(1 for block in blocks if block["type"] == "image"),
                "tableCount": sum(1 for block in blocks if block["type"] == "table"),
                "pageCount": sum(1 for block in blocks if block["type"] == "pdf-page"),
            }
        )

    rows = xlsx_rows(DOCS / "游戏经历.xlsx")
    works.append(
        {
            "category": "experience",
            "slug": "game-experience-table",
            "title": "游戏经历完整表",
            "kind": "Excel 表格 / 游玩履历",
            "format": "xlsx",
            "tags": ["移动端", "主机/PC", "长线体验", "跨类型"],
            "source": "docs/游戏经历.xlsx",
            "image": None,
            "blocks": [],
            "paragraphs": [],
            "table": {
                "headers": ["平台", "游戏类型", "游戏名称", "游戏时长/游戏成就"],
                "rows": rows,
            },
            "paragraphCount": len(rows),
            "imageCount": 0,
            "tableCount": 1,
            "pageCount": 0,
        }
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps({"categories": categories, "works": works}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
